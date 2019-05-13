from openpyxl import load_workbook
import xlsxwriter
import pandas as pd
from dsp.save_data import get_output_fpath, save_param_info


class SaveRawData:
    def __init__(self, filepath, params, result_dict, epr_df, rv_dict,
                 version_info, is_dev):
        self.filepath = filepath
        self.params = params
        self.epr_df = epr_df
        self.rv_dict = rv_dict
        self.res_dict = result_dict
        self.version_info = version_info
        self.is_dev = is_dev

    def export_raw_datasheet(self):
        self.fpath = get_output_fpath(self.filepath, 'raw_datasheet_')
        with xlsxwriter.Workbook(self.fpath) as writer:
            save_param_info(writer, self.params, self.res_dict,
                            self.version_info, is_dev=self.is_dev)
        self.save_res_data()

    def save_res_data(self):
        with pd.ExcelWriter(self.fpath, engine='openpyxl') as writer:
            book = load_workbook(self.fpath)
            writer.book = book
            temp_keys = list(self.params['used_temp'].keys())
            dye_li = []
            for d in self.params['dye_li']:
                if self.params['used_temp'][temp_keys[0]][d] or\
                        self.params['used_temp'][temp_keys[1]][d]:
                    dye_li.append(d)
            for dye in dye_li:
                data_table_li = []
                for i, txt in enumerate(['End RFU (L)', 'End RFU (H)']):
                    data_table_li.append(
                        pd.Series([
                            round(absd[-1], 4) for absd in self.res_dict[
                                'ABSD_array'][temp_keys[i]][dye].values()],
                                  index=self.res_dict[
                                      'ABSD_array'][temp_keys[i]][dye].keys(),
                                  name=txt)
                        )
                for i, txt in enumerate(['RV (L)', 'RV (H)']):
                    data_table_li.append(
                        pd.Series([round(rv, 4) for rv in self.rv_dict[
                            temp_keys[i]][dye].values()], index=self.rv_dict[
                                temp_keys[i]][dye].keys(), name=txt)
                        )
                data_table_df = pd.concat(data_table_li, axis=1, sort=True)
                data_table_df.insert(
                    4, 'EPR',
                    data_table_df['End RFU (L)']/data_table_df['End RFU (H)'])
                data_table_df.fillna('')
                data_table_df.to_excel(writer, sheet_name=dye)
