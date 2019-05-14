import xlsxwriter
from openpyxl import load_workbook
import pandas as pd
import os
import pathlib
import datetime
import copy


def _templess(params, name, key, ws, row):
    col = 0
    ws.write(row, col, name)
    for dye in params['dye_li']:
        col += 1
        ws.write(row, col, params[key][dye])


def _temp(params, name, key, ws, row):
    col = 0
    ws.write(row, col, name)
    for temp in ['Low Temp', 'High Temp']:
        for dye in params['dye_li']:
            col += 1
            ws.write(row, col, params[key][temp][dye])
        col = 0
        row += 1


def write_dye(ws, col, rw, params):
    for dye in params['dye_li']:
        ws.write(rw, col, dye)
        col += 1


def write_pos_ratio(ws, rw, col, params, result_dict):
    write_dye(ws, col+1, rw, params)
    rw += 1
    c = copy.deepcopy(col)
    ct_res = result_dict['Ct_array_final']
    sum_pos_num = 0
    sum_total = 0
    for temp in ct_res.keys():
        ws.write(rw, c, temp)
        c += 1
        for dye in params['dye_li']:
            if len(result_dict['result'][temp][dye]) != 0:
                pos_num = 0
                for well in ct_res[temp][dye].keys():
                    ct_well = ct_res[temp][dye][well]
                    if not isinstance(ct_well, str) and\
                            ct_well > 0 and ct_well < 45:
                        pos_num += 1
                total = len(result_dict['result'][temp][dye])
                percent = round(pos_num/total*100, 2)
                text = str(pos_num) + '/' + str(total) + ' (' + str(percent) +\
                    '%)'
                sum_pos_num += pos_num
                sum_total += total
            else:
                text = '-'
            ws.write(rw, c, text)
            c += 1
        rw += 1
        c = col
    rw += 2
    ws.write(rw, col, 'Total')
    percent = round(sum_pos_num/sum_total*100, 2)
    text = str(sum_pos_num) + '/' + str(sum_total) + ' (' + str(percent) + '%)'
    ws.write(rw, col+1, text)


def save_param_info(writer, params, result_dict, version_info, is_dev=False):
    ws = writer.add_worksheet('parameters')
    row = 0
    if params['is_dev']:
        ws.write(row, 2, 'Latest baseline filters enabled')
    row += 1
    if is_dev:
        ws.write(row, 0, 'Data Scale')
        ws.write(row, 1, params['ds'])
        row += 1
        for i, txt in enumerate([
                'Jumping', 'LinearFilter Type', 'Unimportant', 'Quick Amp',
                'Baseline']):
            ws.write(row, i+1, txt)
        row += 1
        ws.write(row, 0, 'Algorithm Version')
        for i, txt in enumerate(params['ver_array'][3:]):
            ws.write(row, i+1, txt)
        row += 1
    if params['BPN_switch']:
        ws.write(row, 0, 'BPN ON')
    else:
        ws.write(row, 0, 'BPN OFF')
    if params['gib1_switch']:
        ws.write(row, 1, 'GIBI ON')
    elif params['gi_switch']:
        ws.write(row, 1, 'GIV ON')
    elif params['gip']:
        ws.write(row, 1, 'GIP ON')
    else:
        ws.write(row, 1, '제품별 Default')

    if params['absd_corrector_ver']['sample'] and\
            params['absd_corrector_ver']['PC']:
        ws.write(row, 2, '(OLD) 과도한 CR에 대한 보정')
    elif params['absd_corrector_ver']['PC']:
        ws.write(row, 2, '(OLD) 과도한 CR에 대한 보정 - PC only')
    elif params['absd_corrector_ver']['sample']:
        ws.write(row, 2, '(OLD) 과도한 CR에 대한 보정 - sample only')

    if params['asymm_sigfit_ver']['sample'] and\
            params['asymm_sigfit_ver']['PC']:
        ws.write(row, 3, '(OLD) Ct값 보정')
    elif params['asymm_sigfit_ver']['PC']:
        ws.write(row, 3, '(OLD) Ct값 보정 - PC only')
    elif params['asymm_sigfit_ver']['sample']:
        ws.write(row, 3, '(OLD) Ct값 보정 - sample only')

    row += 1
    ws.write(row, 0, 'BPN cycle')
    ws.write(row, 1, params['BPN_used_cycle'][0])
    ws.write(row, 2, params['BPN_used_cycle'][-1])
    row += 2
    write_dye(ws, 1, row, params)
    row += 1
    _templess(params, 'PMC', 'PMC', ws, row)
    row += 1
    _templess(params, 'CR', 'CR', ws, row)
    row += 1
    _templess(params, 'CR (PC)', 'CR_pc', ws, row)
    row += 1
    _templess(params, 'SFC', 'SFC', ws, row)
    row += 1
    _templess(params, 'SFC (PC)', 'SFC_pc', ws, row)
    row += 1
    _templess(params, 'MFC', 'MFC', ws, row)
    row += 1
    _templess(params, 'MFC (PC)', 'MFC_pc', ws, row)
    row += 1
    _templess(params, 'DRFU2', 'PTP', ws, row)
    row += 1
    _templess(params, 'DRFU3', 'drfu3', ws, row)

    row += 3
    _temp(params, 'BPN RV', 'BPN_rv', ws, row)
    row += 2
    _temp(params, 'DRFU', 'DRFU_array', ws, row)
    row += 2
    _temp(params, 'Thrd', 'Thrd_array', ws, row)
    row += 2
    _temp(params, 'RPC', 'RPC_array', ws, row)
    row += 2
    _temp(params, 'RC', 'RC_array', ws, row)
    row += 2
    _temp(params, 'dfM', 'dfM_array', ws, row)
    row += 2
    _temp(params, 'dfC', 'dfC_array', ws, row)
    row += 2
    _temp(params, 'Used Temperature', 'used_temp', ws, row)
    row += 2
    _temp(params, 'Cut Off', 'cut', ws, row)
    row += 2
    _temp(params, 'Cut Off (NC)', 'cut_nc', ws, row)

    row += 3
    ws.write(row, 0, 'Positive Control (PC)')
    row += 1
    _temp(params, 'DRFU for PC', 'drfu_array_pc', ws, row)
    row += 2
    _temp(params, 'Thrd for PC', 'Thrd_array_pc', ws, row)
    row += 2
    _temp(params, 'RPC for PC', 'rpc_array_pc', ws, row)
    row += 2
    _temp(params, 'RC for PC', 'rc_array_pc', ws, row)
    row += 2
    _temp(params, 'dfM for PC', 'dfM_array_pc', ws, row)
    row += 2
    _temp(params, 'dfC for PC', 'dfC_array_pc', ws, row)

    crosstalk_li = params['crosstalk']
    if crosstalk_li:
        row += 3
        ws.write(row, 0, 'Crosstalk')
        row += 1
        ws.write(row, 0, 'target1')
        ws.write(row, 1, 'target2')
        ws.write(row, 2, 'crosstalk drfu')
        ws.write(row, 3, 'crosstalk thrd')
        row += 1
        for crosstalk in crosstalk_li:
            col = 0
            for c in crosstalk:
                ws.write(row, col, c)
                col += 1
            row += 1
    write_pos_ratio(ws, 7, 7, params, result_dict)


def get_output_fpath(filepath, prefix=''):
    _fname = pathlib.PurePath(
        filepath.name.translate(str.maketrans('', '', '[]<>?|*')))
    inputf = filepath.parents[0]/_fname
    save_path = inputf.parents[1]
    fname = prefix + inputf.name.split(' -  ')[0] + inputf.suffix
    if os.path.exists(save_path/fname):
        name, ext = fname.split('.')
        fname = name + datetime.datetime.now().strftime(
            '_%Y%m%d_%H%M%S') + inputf.suffix
    fpath = save_path/fname
    return fpath


def save_data_fn(filepath, params, result_dict, version_info, is_dev):
    fpath = get_output_fpath(filepath)
    with xlsxwriter.Workbook(fpath) as writer:
        save_param_info(writer, params, result_dict, version_info, is_dev)

    with pd.ExcelWriter(fpath, engine='openpyxl') as writer:
        book = load_workbook(fpath)
        writer.book = book
        temp_keys = list(result_dict['result'].keys())
        dye_li = []
        for d in params['dye_li']:
            if params['used_temp'][temp_keys[0]][d] or\
                    params['used_temp'][temp_keys[1]][d]:
                dye_li.append(d)

        for dye in dye_li:
            data_table_li = []
            for key in ['data_process_num', 'Ct_array_final', 'R_p2_array',
                        'R2_array', 'end_rfu', 'df_array']:
                for i in range(2):
                    data_table_li.append(result_dict[key][temp_keys[i]][dye])
            df_li = [pd.DataFrame(
                data_table_li, index=['Low', 'High', 'Ct (L)', 'Ct (H)',
                                      'Rp2 (L)', 'Rp2 (H)', 'R2 (L)', 'R2 (H)',
                                      'RFU (L)', 'RFU (H)', 'df (L)',
                                      'df (H)'])]
            for i, ind in enumerate(
                    [['cff1 (L)', 'cff2 (L)'], ['cff1 (H)', 'cff2 (H)']]):
                df_li.append(
                    pd.DataFrame(result_dict['cff_array'][temp_keys[i]][dye],
                                 index=ind))

            lsr_low = {}
            lsr_high = {}
            for temp in temp_keys:
                param = {}
                for well in result_dict['df_array'][temp][dye].keys():
                    result = result_dict['result_array_orig'][temp][dye][well]
                    if result == -2:
                        param[well] = result_dict[
                            'param_new_array'][temp][dye][well]
                    else:
                        param[well] = result_dict[
                            'param_array'][temp][dye][well]
                    if result <= 0:
                        if 'Low' in temp:
                            lsr_low[well] = result_dict[
                                'LSR_val_array'][temp][dye][well]
                        else:
                            lsr_high[well] = result_dict[
                                'LSR_val_array'][temp][dye][well]
                    else:
                        if 'Low' in temp:
                            lsr_low[well] = 0
                        else:
                            lsr_high[well] = 0
                param_ind = [f'a1 ({temp[0]})', f'a2 ({temp[0]})',
                             f'a3 ({temp[0]})', f'a4 ({temp[0]})']
                df_li.append(pd.DataFrame(param, index=param_ind))
            df_li.append(pd.DataFrame([lsr_low], index=['LSR (L)']))
            df_li.append(pd.DataFrame([lsr_high], index=['LSR (H)']))
            data_table_df = pd.concat(df_li, sort=True).T
            data_table_df.insert(
                10, 'EPR', data_table_df['RFU (L)']/data_table_df['RFU (H)'])
            data_table_df.fillna('')
            data_table_df.to_excel(writer, sheet_name=dye)
